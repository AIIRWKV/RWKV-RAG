# coding=utf-8

from openpyxl import load_workbook

from src.diversefile import AbstractLoader

class ExcelLoader(AbstractLoader):
    """
    按行chunking
    """
    def load(self):
        wb = load_workbook(self.file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = ws.rows #  generator
            try:
                title = next(rows)
            except StopIteration:
                continue

            for row in rows:
                lines = []
                for i, c in enumerate(row):
                    if not c.value:
                        continue
                    t = str(title[i].value) if i < len(title) else ""
                    t += ("：" if t else "") + str(c.value)
                    lines.append(t)
                l = "; ".join(lines)
                if sheet_name.lower().find("sheet") < 0:
                    l += " ——" + sheet_name
                yield l
        wb.close()