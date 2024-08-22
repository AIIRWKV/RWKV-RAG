import os
import io
import subprocess
from typing import Literal


import fitz
from PIL import Image
try:
    import pytesseract
except ImportError:
    pass

class Loader:
    """
    加载器，用来分割本地文件
    """
    def __init__(self, file_path: str, chunk_size:int, chunk_overlap:int=0, remove_linefeed=True):
        """
        :param file_path: 文件路径或者目录地址
        :param chunk_size: 块大小
        :param chunk_overlap: 块重叠区间大小
        :param remove_linefeed: 是否去掉换行符
        """
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.remove_linefeed = remove_linefeed
        self._files = []
        self.output_files = []
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File or Directory not found: {file_path}")
        if os.path.isdir(file_path):
            for _file in os.listdir(file_path):
                _file_path = os.path.join(file_path, _file)
                if os.path.isfile(_file_path):
                    self._files.append(_file_path)
        else:
            self._files = [self.file_path]

    def load_txt(self, file_path: str, split_type=None, lang=None):
        """
        加载txt文件
        """
        chunk_overlap = self.chunk_overlap
        chunk_size = self.chunk_size
        batch_size = 4086
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            current_txt = ''
            while 1:
                txt = f.read(batch_size)
                if not txt:
                    break
                txt = txt.strip().replace("\r\n", "").replace("\n", "") if self.remove_linefeed else txt.strip()
                current_txt += txt
                if len(current_txt) >= self.chunk_size:
                    yield current_txt[:chunk_size]
                    current_txt = current_txt[chunk_size - chunk_overlap:]
            if current_txt:
                yield current_txt

    @staticmethod
    def extract_text_from_image(image_data, language='chi_sim'):
        # 将图像数据转换为 PIL Image 对象
        image = Image.open(io.BytesIO(image_data))
        # 使用 Tesseract 进行 OCR 识别
        text = pytesseract.image_to_string(image, language)  # 设置识别语言
        return text

    def load_pdf(self, file_path: str, split_type=None, lang='中文'):
        tesseract_lang = {'中文': 'chi_sim', '英文': 'eng', '日语': 'jpn'}
        if lang not in tesseract_lang:
            raise ValueError(f"Language not supported: {lang} when load PDF file")
        language = tesseract_lang[lang]
        try:
            result = subprocess.run(['tesseract', '--list-langs'], capture_output=True, text=True)
            if result.returncode != 0:
                raise ValueError("Tesseract not found. Please install Tesseract and add it to your PATH.")
        except FileNotFoundError:
            raise ValueError("Tesseract not found. Please install Tesseract and add it to your PATH.")
        stdout = result.stdout
        stdout_list = stdout.split('\n')
        has_language = False
        for line in stdout_list:
            line = line.strip()
            if language == line:
                has_language = True
                break
        if not has_language:
            raise ValueError(f"tessdata failed loading language '{lang}' Tesseract couldn\'t load any languages! Could not initialize tesseract.")
        with fitz.open(file_path) as doc:
            current_txt = ''
            chunk_overlap = self.chunk_overlap
            chunk_size = self.chunk_size
            for page in doc:
                txt = page.get_text()
                if txt:
                    txt = txt.strip().replace("\r\n", "").replace("\n", "") if self.remove_linefeed else txt.strip()
                    current_txt += txt
                else:
                    imgs = page.get_images(full=True)
                    for img in imgs:
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        image_data = base_image["image"]
                        # 调用 OCR 函数提取文本
                        try:
                            txt = self.extract_text_from_image(image_data, language=language)
                            txt = txt.strip().replace("\r\n", "").replace("\n", "") if self.remove_linefeed else txt.strip()
                        except:
                            continue
                        current_txt += txt

                while len(current_txt) >= chunk_size:
                    yield current_txt[:chunk_size]
                    current_txt = current_txt[chunk_size - chunk_overlap:]

            if current_txt:
                yield current_txt


    def load_xlsx(self, file_path: str, split_type:  Literal['cell', 'row'] ='cell', lang=None):
        """
        Load excel
        :param file_path: excel file path
        :param split_type: 'cell', 'row'
                           cell: 按excel的单元格来分割，此时chunk_size, chunk_overlap参数无效
                           row: 按行读取，并用chunk_size, chunk_overlap的方式分割

        """
        from openpyxl import load_workbook
        # 加载工作簿
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        chunk_overlap = self.chunk_overlap
        chunk_size = self.chunk_size
        if split_type == 'cell':
            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                # 遍历每一行
                for rows in worksheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始
                    for row in rows:
                        if not row:
                            continue
                        if isinstance(row, str):
                            row = row.strip().replace("\r\n", "").replace("\n", "") \
                                if self.remove_linefeed else row
                            yield row
                        else:
                            yield str(row)

        elif split_type == 'row':
            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                current_txt = ''
                # 遍历每一行
                for rows in worksheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始
                    line = []
                    for row in rows:
                        if not row:
                            continue
                        if isinstance(row, str):
                            line.append(row.strip().replace("\r\n", "").replace("\n", "")
                                        if self.remove_linefeed else row)
                        else:
                            line.append(str(row))
                    current_txt += ' '.join(line)

                    while len(current_txt) >= chunk_size:
                        yield current_txt[:chunk_size]
                        current_txt = current_txt[chunk_size - chunk_overlap:]


        else:
            raise ValueError(f"Excel Unsupported split_type: {split_type}")


        workbook.close()

    load_xls = load_xlsx

    def load_and_split_file(self, output_dir: str,  split_type='cell', lang='中文'):
        for path in self._files:
            base_filename, file_ext = os.path.splitext(os.path.basename(path))
            file_ext = file_ext.lstrip('.')
            func = getattr(self, f'load_{file_ext}', None)
            lang = lang
            if func and callable(func):
                output_file = os.path.join(output_dir, f"{base_filename}_chunks.txt")
                self.output_files.append(output_file)
                with open(output_file, 'w', encoding='utf-8') as f:
                    for item in func(path, split_type=split_type, lang=lang):
                        if item:
                            yield item
                            f.write(item)
                            f.write('\n')


